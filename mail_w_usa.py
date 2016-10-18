# Create Sendout worksheet for Sendout Cards

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

wb = load_workbook('nool.xlsx')
wb_sheets =  wb.get_sheet_names()

ws_export = wb["Export"]
ws_sendout = wb["Sendout-US"]

def parse_names (i,j):
    # mail_name contains the raw column from rebo gateway
    mail_name = ws_export.cell(row=i, column=18)
    if "," in mail_name.value:
        lastname = mail_name.value.split(",",1)[0]
        raw_first_names = mail_name.value.split(",",1)[1]
    else:
        lastname = mail_name.value
        raw_first_names = ""

    if "&" in raw_first_names:
        spouse_1 = raw_first_names.split(" & ",1)[0]
        spouse_2 = raw_first_names.split(" & ",1)[1]
        if " " in spouse_1:
            spouse_1_name = spouse_1.split(" ")[0]
            spouse_1_middle = spouse_1.split(" ")[1]
        else:
            spouse_1_name = spouse_1.split(" ")[0]
            spouse_1_middle = ""
        if " " in spouse_2:
            spouse_2_name = spouse_2.split(" ")[0]
            spouse_2_middle = spouse_2.split(" ")[1]
        else:
            spouse_2_name = spouse_2.split(" ")[0]
            spouse_2_middle = ""
    else:
        spouse_1 = raw_first_names.split(" & ",1)[0]
        spouse_2 = ""
        spouse_2_name = ""
        spouse_2_middle = ""
        if " " in spouse_1:
            spouse_1_name = spouse_1.split(" ")[0]
            spouse_1_middle = spouse_1.split(" ")[1]
        else:
            spouse_1_name = spouse_1.split(" ")[0]
            spouse_1_middle = ""
    # first_name
    ws_sendout.cell(row=j, column=1).value = spouse_1_name
    # Spouse
    ws_sendout.cell(row=j, column=17).value = spouse_2_name
    ws_sendout.cell(row=j, column=2).value = lastname
    return

# Initialize the output worksheet
ws_sendout.cell(row=1, column=1).value = "first_name"
ws_sendout.cell(row=1, column=2).value = "last_name"
ws_sendout.cell(row=1, column=3).value = "address1"
ws_sendout.cell(row=1, column=4).value = "address2"
ws_sendout.cell(row=1, column=5).value = "city"
ws_sendout.cell(row=1, column=6).value = "state"
ws_sendout.cell(row=1, column=7).value = "postal_code"
ws_sendout.cell(row=1, column=8).value = "country"
ws_sendout.cell(row=1, column=9).value = "company"
ws_sendout.cell(row=1, column=10).value = "email_address"
ws_sendout.cell(row=1, column=11).value = "home_phone"
ws_sendout.cell(row=1, column=12).value = "work_phone"
ws_sendout.cell(row=1, column=13).value = "cell_number"
ws_sendout.cell(row=1, column=14).value = "fax_number"
ws_sendout.cell(row=1, column=15).value = "birthday"
ws_sendout.cell(row=1, column=16).value = "anniversary"
ws_sendout.cell(row=1, column=17).value = "spouse"
ws_sendout.cell(row=1, column=18).value = "spouse_birthday"
ws_sendout.cell(row=1, column=19).value = "Group"

tot_rows = ws_export.max_row + 1
j = 1
for i in range(2, tot_rows):
    country = ws_export.cell(row=i, column=24).value
    if country == "US":
        j = j + 1;
        parse_names(i,j);
        # Mailing address city and/or state are not blank send cards here
        # Mailing address (address1)
        ws_sendout.cell(row=j, column=3).value = ws_export.cell(row=i, column=19).value;
        # Mailing address (address2)
        ws_sendout.cell(row=j, column=4).value = ws_export.cell(row=i, column=20).value;
        # Mailing address (city)
        ws_sendout.cell(row=j, column=5).value = ws_export.cell(row=i, column=21).value;
        # Mailing address (state)
        ws_sendout.cell(row=j, column=6).value = ws_export.cell(row=i, column=22).value;
        # Mailing address (postal_code)
        ws_sendout.cell(row=j, column=7).value = ws_export.cell(row=i, column=23).value;
        # Mailing address (country)
        ws_sendout.cell(row=j, column=8).value = "USA";

tot_entries = j - 1
print "Created Sendout-US: ", tot_entries
# Save the workbook
wb.save('nool_sendout.xlsx')
